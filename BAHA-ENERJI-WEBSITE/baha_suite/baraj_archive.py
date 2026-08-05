"""EPİAŞ aktif doluluk kayıtlarını mevcut Excel arşivine güvenle ekler."""

from __future__ import annotations

import math
import os
import threading
from copy import copy
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4


_ARCHIVE_WRITE_LOCK = threading.Lock()


@dataclass(frozen=True)
class ArchiveUpdateResult:
    updated: bool
    added_dates: tuple[str, ...] = ()
    added_rows: int = 0
    reason: str = ""

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


def _day(value: Any) -> date | None:
    text = str(value or "").strip()[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _name(value: Any) -> str:
    return str(value or "").strip()


def _name_key(value: Any) -> str:
    return _name(value).casefold()


def _number(value: Any) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _normalized_days(
    items: Iterable[dict[str, Any]],
    *,
    latest_allowed_date: date,
) -> dict[date, list[dict[str, Any]]]:
    by_day: dict[date, dict[str, dict[str, Any]]] = {}
    for item in items:
        selected_day = _day(item.get("date"))
        dam = _name(item.get("dam") or item.get("damName"))
        basin = _name(item.get("basin") or item.get("basinName"))
        fullness = _number(item.get("activeFullnessAmount"))
        if (
            selected_day is None
            or selected_day > latest_allowed_date
            or not dam
            or not basin
            or fullness is None
        ):
            continue
        by_day.setdefault(selected_day, {})[_name_key(dam)] = {
            "date": selected_day,
            "basin": basin,
            "dam": dam,
            "activeFullnessAmount": fullness,
        }
    return {selected_day: list(rows.values()) for selected_day, rows in by_day.items()}


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _average(values: Iterable[Any]) -> float | None:
    numbers = [number for value in values if (number := _number(value)) is not None]
    return sum(numbers) / len(numbers) if numbers else None


def _existing_raw_dates(worksheet: Any) -> set[date]:
    result: set[date] = set()
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        value = row[0]
        if isinstance(value, datetime):
            result.add(value.date())
        elif isinstance(value, date):
            result.add(value)
        elif parsed := _day(value):
            result.add(parsed)
    return result


def _find_cell(worksheet: Any, column: int, expected: str) -> int:
    expected_key = expected.casefold()
    for row in range(1, worksheet.max_row + 1):
        if _name(worksheet.cell(row, column).value).casefold() == expected_key:
            return row
    raise ValueError(f"Excel arşivinde '{expected}' hücresi bulunamadı.")


def _pivot_date_columns(worksheet: Any, header_row: int) -> list[int]:
    result: list[int] = []
    for column in range(2, worksheet.max_column + 1):
        value = worksheet.cell(header_row, column).value
        if isinstance(value, (datetime, date)) or _day(value):
            result.append(column)
    return result


def _append_raw_rows(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    template_row = max(2, worksheet.max_row)
    for item in rows:
        target_row = worksheet.max_row + 1
        for column in range(1, max(5, worksheet.max_column) + 1):
            _copy_cell_style(
                worksheet.cell(template_row, column),
                worksheet.cell(target_row, column),
            )
        worksheet.cell(target_row, 1).value = datetime.combine(item["date"], time.min)
        worksheet.cell(target_row, 2).value = item["basin"]
        worksheet.cell(target_row, 3).value = item["dam"]
        worksheet.cell(target_row, 4).value = item["activeFullnessAmount"]
        template_row = target_row


def _append_pivot_day(worksheet: Any, selected_day: date, rows: list[dict[str, Any]]) -> None:
    header_row = _find_cell(worksheet, 1, "Barajlar")
    total_row = _find_cell(worksheet, 1, "Genel Ortalama")
    grand_average_column = next(
        (
            column
            for column in range(2, worksheet.max_column + 1)
            if _name(worksheet.cell(header_row, column).value).casefold()
            == "genel ortalama"
        ),
        None,
    )
    if grand_average_column is None:
        raise ValueError("Pivot sekmesinde Genel Ortalama sütunu bulunamadı.")

    dam_rows = {
        _name_key(worksheet.cell(row, 1).value): row
        for row in range(header_row + 1, total_row)
        if _name(worksheet.cell(row, 1).value)
    }
    missing_dams = [item for item in rows if _name_key(item["dam"]) not in dam_rows]
    for item in sorted(missing_dams, key=lambda value: _name_key(value["dam"])):
        worksheet.insert_rows(total_row)
        source_row = max(header_row + 1, total_row - 1)
        for column in range(1, worksheet.max_column + 1):
            _copy_cell_style(
                worksheet.cell(source_row, column),
                worksheet.cell(total_row, column),
            )
            worksheet.cell(total_row, column).value = None
        worksheet.cell(total_row, 1).value = item["dam"]
        dam_rows[_name_key(item["dam"])] = total_row
        total_row += 1

    worksheet.insert_cols(grand_average_column)
    style_source_column = max(2, grand_average_column - 1)
    for row in range(1, worksheet.max_row + 1):
        _copy_cell_style(
            worksheet.cell(row, style_source_column),
            worksheet.cell(row, grand_average_column),
        )
        worksheet.cell(row, grand_average_column).value = None
    worksheet.cell(header_row, grand_average_column).value = datetime.combine(
        selected_day,
        time.min,
    )

    value_by_dam = {
        _name_key(item["dam"]): item["activeFullnessAmount"] for item in rows
    }
    for dam_key, row in dam_rows.items():
        worksheet.cell(row, grand_average_column).value = value_by_dam.get(dam_key)
    worksheet.cell(total_row, grand_average_column).value = _average(value_by_dam.values())

    shifted_grand_average_column = grand_average_column + 1
    worksheet.cell(header_row, shifted_grand_average_column).value = "Genel Ortalama"
    date_columns = _pivot_date_columns(worksheet, header_row)
    all_values: list[float] = []
    for row in range(header_row + 1, total_row):
        values = [worksheet.cell(row, column).value for column in date_columns]
        worksheet.cell(row, shifted_grand_average_column).value = _average(values)
        all_values.extend(
            number for value in values if (number := _number(value)) is not None
        )
    worksheet.cell(total_row, shifted_grand_average_column).value = _average(all_values)


def _refresh_pivot_metadata(workbook: Any, raw_sheet: Any, pivot_sheet: Any) -> None:
    from openpyxl.utils import get_column_letter

    for pivot in getattr(pivot_sheet, "_pivots", ()):
        cache = getattr(pivot, "cache", None)
        cache_source = getattr(cache, "cacheSource", None)
        worksheet_source = getattr(cache_source, "worksheetSource", None)
        if worksheet_source is not None:
            worksheet_source.ref = f"A1:D{raw_sheet.max_row}"
            worksheet_source.sheet = raw_sheet.title
        if cache is not None:
            cache.refreshOnLoad = True
            cache.enableRefresh = True
        if getattr(pivot, "location", None) is not None:
            pivot.location.ref = (
                f"A3:{get_column_letter(pivot_sheet.max_column)}{pivot_sheet.max_row}"
            )
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True


def append_active_fullness_days(
    workbook_path: str | Path,
    items: Iterable[dict[str, Any]],
    *,
    minimum_records: int = 50,
    latest_allowed_date: date | None = None,
) -> ArchiveUpdateResult:
    """Arşivden yeni ve yeterince dolu EPİAŞ günlerini kronolojik olarak ekler.

    Mevcut günler Excel otoritesi sayılır ve asla değiştirilmez. EPİAŞ cevabındaki
    yalnızca en yeni gün, arşiv tarihinden ilerideyse kabul edilir. Eski eksik günler
    sonradan geri doldurulmaz; böylece tarih sırası ve kilitli geçmiş korunur.
    """

    path = Path(workbook_path)
    if not path.is_file():
        return ArchiveUpdateResult(False, reason="Excel arşivi bulunamadı.")
    normalized = _normalized_days(
        items,
        latest_allowed_date=latest_allowed_date or date.today(),
    )
    if not normalized:
        return ArchiveUpdateResult(False, reason="Geçerli EPİAŞ baraj kaydı yok.")

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("Excel arşivi için openpyxl kurulmalıdır.") from exc

    with _ARCHIVE_WRITE_LOCK:
        workbook = openpyxl.load_workbook(path)
        try:
            raw_sheet = workbook["Aktif Doluluk"]
            pivot_sheet = workbook["Pivot"]
            existing_dates = _existing_raw_dates(raw_sheet)
            latest_existing = max(existing_dates) if existing_dates else date.min
            # Aktif doluluk servisi arşiv geri doldurma amacıyla kullanılmaz.
            # Cevap birden fazla gün taşısa bile yalnızca en yeni yayımlanan gün
            # eklenir; geçmiş Excel günleri dokunulmaz arşiv olarak kalır.
            newest_published_day = max(normalized)
            pending = (
                [newest_published_day]
                if newest_published_day > latest_existing
                and newest_published_day not in existing_dates
                and len(normalized[newest_published_day])
                >= max(1, minimum_records)
                else []
            )
            if not pending:
                return ArchiveUpdateResult(
                    False,
                    reason=(
                        "Yeni tam gün yok; mevcut Excel kayıtları korunuyor."
                    ),
                )

            added_rows = 0
            for selected_day in pending:
                rows = normalized[selected_day]
                _append_raw_rows(raw_sheet, rows)
                _append_pivot_day(pivot_sheet, selected_day, rows)
                added_rows += len(rows)
            _refresh_pivot_metadata(workbook, raw_sheet, pivot_sheet)

            temporary_path = path.with_name(f".{path.name}.{uuid4().hex}.tmp.xlsx")
            try:
                workbook.save(temporary_path)
                os.replace(temporary_path, path)
            finally:
                if temporary_path.exists():
                    temporary_path.unlink()
            return ArchiveUpdateResult(
                True,
                added_dates=tuple(selected_day.isoformat() for selected_day in pending),
                added_rows=added_rows,
                reason="EPİAŞ günleri Excel arşivine eklendi.",
            )
        finally:
            workbook.close()
